#! /usr/bin/python3
# -*- coding: utf-8 -*-

"""
@Author: Kobayasi
@File: excel.py
@Time: 2022/8/11-16:40
"""

import copy
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from typing import List


class CreateExcel:

    def __init__(self, path: str):
        """
        保存文件的路径 xlsx格式
        :param path:
        """
        self.path = path
        self.workbook = openpyxl.Workbook()

    def insert(self, sheet_name: List[str], sheet_title: List[str], sheet_data: List[List[list]]):
        """
        插入数据到excel
        :param sheet_name: [name1, name2]
        :param sheet_title: [title1, title2]
        :param sheet_data:
        [
            [
                [name1-value, name1-value],
                [name1-value, name1-value],
            ],
            [
                [name2-value, name2-value],
                [name2-value, name2-value],
            ],
        ]
        :return:
        """
        if len(sheet_name) != len(sheet_data):
            raise ValueError(f'sheet列表与data列表对应不一致: name len {len(sheet_name)}, data len {len(sheet_data)}')

        sheet_num = 0
        for sheet, data in zip(sheet_name, sheet_data):
            worksheet = self.workbook.create_sheet(title=sheet, index=sheet_num)
            # 写入表头
            for num in range(len(sheet_title)):
                worksheet.cell(row=1, column=num + 1).value = sheet_title[num]

            # 写入数据
            for row_num in range(len(data)):
                for column_num in range(len(data[row_num])):
                    value = data[row_num][column_num]
                    if isinstance(value, dict):
                        value = json.dumps(value, indent=4, ensure_ascii=False, sort_keys=False)
                        worksheet.cell(row=row_num + 2, column=column_num + 1).alignment = Alignment(wrap_text=True)
                    worksheet.cell(row=row_num + 2, column=column_num + 1).value = value

            sheet_num += 1

        self.workbook.save(self.path)


class ReadExcel:

    def __init__(self, path: str, case_id: int):
        self.wb = load_workbook(path)
        self.case_id = case_id

    async def read(self):
        sheet = self.wb.active
        # 把excel数据先读出来
        data_list = []
        url = None
        da = None
        for column in range(1, sheet.max_column + 1):
            row_list = []
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=column).value
                if row == 1:
                    if value:
                        url = value
                        row_list.append(url)
                    else:
                        row_list.append(url)
                elif row == 2:
                    if value:
                        da = value
                        row_list.append(da)
                    else:
                        row_list.append(da)
                else:
                    row_list.append(value)
            data_list.append(row_list)
        self.wb.close()

        # 处理数据，一行数据集一套
        gather_name = []
        new_data_list = [{} for _ in range(len(data_list[0][4:]))]
        num = 0
        for column in data_list:
            if num != 0:
                for x in range(len(column)):
                    if x == 0:
                        for i in range(len(new_data_list)):
                            if not new_data_list[i].get(column[x]):
                                new_data_list[i][column[x]] = {}

                    elif x == 1:
                        for i in range(len(new_data_list)):
                            if not new_data_list[i][column[0]].get(column[1]):
                                new_data_list[i][column[0]][column[1]] = {}
                    elif x == 2:
                        for i in range(len(new_data_list)):
                            if not new_data_list[i][column[0]][column[1]].get(column[2]):
                                new_data_list[i][column[0]][column[1]][column[2]] = column[4 + i]
            else:
                gather_name = column[4:]

            num += 1

        new_data = []
        num = 1
        for name, gather_data in zip(gather_name, new_data_list):

            for k, v in gather_data.items():
                new_dict = {}
                number, url = k.split("#")
                new_dict['case_id'] = self.case_id
                new_dict['suite'] = num
                new_dict['name'] = name
                new_dict['number'] = int(number)
                new_dict['path'] = url
                new_dict = {**new_dict, **v}
                new_data.append(new_dict)
            num += 1

        return new_data
