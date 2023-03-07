#! /usr/bin/python3
# -*- coding: utf-8 -*-

"""
@Author: Kobayasi
@File: case_data_gather.py
@Time: 2023/3/6-14:46
"""

from typing import Any
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class CaseDataGather:

    def __init__(self):

        self.color1 = ['DAEEF3', 'B7DEE8', '92CDDC', '31869B']
        self.color2 = ['FDE9D9', 'FCD5B4', 'FABF8F', 'E26B0A']

    async def data_gather(self, case_data, path: str):
        """
        处理测试数据集的内容
        :param case_data:
        :param path:
        :return:
        """
        # 处理需要做数据集的参数
        case_gather = []
        for data in case_data:
            # params的参数数据
            params = self._header_data_list(self._header_data(data.params))
            # data的参数数据
            data_ = self._header_data_list(self._header_data(data.data))
            # check的数据
            check = self._header_data_list(self._header_data(data.check))
            if params or data_:
                case_gather.append({
                    'url': f"{data.number}#{data.path}",
                    'params': params,
                    'data': data_,
                    'check': check

                })
        # 输出表格
        await self._excel(path=path, case_gather=case_gather)

    # @staticmethod
    async def _excel(self, path: str, case_gather: list):
        """
        输出excel
        :param case_gather:
        :param path
        :return:
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.freeze_panes = 'B4'
        # align = Alignment(horizontal='center', vertical='center')
        start_column = 2
        end_column = 0
        for data in case_gather:
            params_len = len(data['params'])
            data_len = len(data['data'])
            check_len = len(data['check'])
            url_len = params_len + data_len + check_len

            end_column += url_len
            # 合并url的行
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=start_column,
                end_column=end_column + 1
            )
            self._sheet_coll(
                sheet=sheet,
                row=1,
                col=start_column,
                value=data['url'],
                fill=PatternFill('solid', fgColor=self.color2[0] if start_column % 2 else self.color1[0])
            )
            self._sheet_coll(sheet=sheet, row=1, col=1, value='URL')
            if data['params']:
                sheet.merge_cells(
                    start_row=2,
                    end_row=2,
                    start_column=start_column,
                    end_column=end_column + 1 - (data_len + check_len)
                )
                self._sheet_coll(
                    sheet=sheet,
                    row=2,
                    col=start_column,
                    value='params',
                    fill=PatternFill('solid', fgColor=self.color2[1] if start_column % 2 else self.color1[1])
                )
                self._sheet_coll(sheet=sheet, row=2, col=1, value='分类')
                for x in range(len(data['params'])):
                    for y in range(len(data['params'][x])):
                        self._sheet_coll(
                            sheet=sheet,
                            row=3 + y,
                            col=start_column + x,
                            value=data['params'][x][y]
                        )
            if data['data']:
                sheet.merge_cells(
                    start_row=2,
                    end_row=2,
                    start_column=start_column + params_len,
                    end_column=end_column + 1 - check_len
                )
                self._sheet_coll(
                    sheet=sheet,
                    row=2,
                    col=start_column + params_len,
                    value='data',
                    fill=PatternFill('solid', fgColor=self.color2[2] if start_column % 2 else self.color1[2])
                )
                for x in range(len(data['data'])):
                    for y in range(len(data['data'][x])):
                        self._sheet_coll(
                            sheet=sheet,
                            row=3 + y,
                            col=start_column + params_len + x,
                            value=data['data'][x][y]
                        )

            if data['check']:
                sheet.merge_cells(
                    start_row=2,
                    end_row=2,
                    start_column=start_column + params_len + data_len,
                    end_column=end_column + 1
                )
                self._sheet_coll(
                    sheet=sheet,
                    row=2,
                    col=start_column + params_len + data_len,
                    value='check',
                    fill=PatternFill('solid', fgColor=self.color2[3] if start_column % 2 else self.color1[3])
                )
                self._sheet_coll(sheet=sheet, row=3, col=1, value='参数')
                self._sheet_coll(sheet=sheet, row=4, col=1, value='原始数据')
                self._sheet_coll(sheet=sheet, row=5, col=1, value='数据集-1')
                self._sheet_coll(sheet=sheet, row=6, col=1, value='数据集-2')
                self._sheet_coll(sheet=sheet, row=7, col=1, value='数据集-3')
                for x in range(len(data['check'])):
                    for y in range(len(data['check'][x])):
                        self._sheet_coll(
                            sheet=sheet,
                            row=3 + y,
                            col=start_column + params_len + data_len + x,
                            value=data['check'][x][y]
                        )
            start_column += url_len

        workbook.save(path)

    @staticmethod
    def _header_data(data_json: dict):
        """
        处理数据
        :param data_json:
        :return:
        """
        target = {}

        def inter_of(data):
            for k, v in data.items():
                if not isinstance(v, (list, dict)):
                    if '{{' not in str(v) and "$" not in str(v) and '}}' not in str(v):
                        target[k] = v
                    continue
                if isinstance(v, dict):
                    target[k] = inter_of(v)
                    continue
                if isinstance(v, list):
                    # new_list = []
                    for x in v:
                        inter_of(x)
                    # target[k] = new_list
                    continue
            return target

        return inter_of(data_json)

    @staticmethod
    def _header_data_list(data_json: dict):
        """
        把字典数据处理成list
        :param data_json:
        :return:
        """
        return [[k, v] for k, v in data_json.items()]

    @staticmethod
    def _sheet_coll(
            sheet,
            row: int,
            col: int,
            value: Any,
            align=Alignment(horizontal='center', vertical='center'),
            fill=PatternFill()
    ):
        """
        生成单元格样试
        :param sheet: 表格
        :param row: 行
        :param col: 列
        :param value: 值
        :param align: 居中
        :param fill: 背景填充
        :return:
        """
        sheet.cell(row, col).value = value
        sheet.cell(row, col).alignment = align
        sheet.cell(row, col).fill = fill
