#! /usr/bin/python3
# -*- coding: utf-8 -*-

"""
@Author: Kobayasi
@File: excel.py
@Time: 2022/8/11-16:40
"""

import json
import openpyxl
from openpyxl.styles import Alignment
from typing import List


class CreateExcel:

    def __init__(self, path: str):
        """
        保存文件的路径 xlsx格式
        :param path:
        """
        self.path = path
        self.workbook = openpyxl.Workbook()

    def insert(self, sheet_name: List[str], sheet_title: List[str], sheet_data: List[List[list]]):
        """
        插入数据到excel
        :param sheet_name: [name1, name2]
        :param sheet_title: [title1, title2]
        :param sheet_data:
        [
            [
                [name1-value, name1-value],
                [name1-value, name1-value],
            ],
            [
                [name2-value, name2-value],
                [name2-value, name2-value],
            ],
        ]
        :return:
        """
        if len(sheet_name) != len(sheet_data):
            raise ValueError(f'sheet列表与data列表对应不一致: name len {len(sheet_name)}, data len {len(sheet_data)}')

        sheet_num = 0
        for sheet, data in zip(sheet_name, sheet_data):
            worksheet = self.workbook.create_sheet(title=sheet, index=sheet_num)
            # 写入表头
            for num in range(len(sheet_title)):
                worksheet.cell(row=1, column=num + 1).value = sheet_title[num]

            # 写入数据
            for row_num in range(len(data)):
                for column_num in range(len(data[row_num])):
                    value = data[row_num][column_num]
                    if isinstance(value, dict):
                        value = json.dumps(value, indent=4, ensure_ascii=False, sort_keys=False)
                        worksheet.cell(row=row_num + 2, column=column_num + 1).alignment = Alignment(wrap_text=True)
                    worksheet.cell(row=row_num + 2, column=column_num + 1).value = value

            sheet_num += 1

        self.workbook.save(self.path)
